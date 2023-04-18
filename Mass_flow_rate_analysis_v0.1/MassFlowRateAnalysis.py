'''
Author: Piotr K.
Version: v0.1
Date: 19.04.2023

The class is intended to read files from experimental measurement and generate report into Excel file.
The measurements files have to have defined format and structure, see database folder. 
What's more whole measurement campaign is divided for multiple measurement cases, and each case consist of a few files.   

Measurement case - a set of measurements which consist with some number of files
Measurement file - a result of single measurement

'''
import pandas as pd
import numpy as np
import os
import json
import string
import itertools
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl import Workbook
from openpyxl.chart import (
    ScatterChart,
    Series,
    Reference
)

class mass_flow_rate_analysis:
    def __init__(self,idx_df_weight = 4,idx_df_pressure = 3,window_size=100):
        self.idx_df_weight = idx_df_weight
        self.idx_df_pressure = idx_df_pressure
        self.window_size = window_size 
        # self.horizontal_offset=horizontal_offset

    def open_dir_list(self,json_file_path):
            '''Open dictionary file (*.json) with a list of measurement cases directory'''
            f=open(json_file_path)
            direcory_list=json.load(f)
            f.close()
            return direcory_list

    def get_files_from_dir(self,folder_path):
        '''Prepare a list of files in provided directory'''
        files = list()
        for file in os.listdir(folder_path):
            if os.path.isfile(os.path.join(folder_path, file)):
                files.append(file)
        return files
    
    def read_parameters_from_header(self,file_path):
        '''Read basic parameters from measurement file header'''
        df_head=pd.read_csv(file_path,sep='\t',header=[0])
        date,hour,data_rate=df_head.columns[0],df_head.columns[1],df_head.columns[3]
        data_rate=int(data_rate)
        return date,hour,data_rate
    
    def join_tuple(self,tuple,comma_separator=True):
        '''If data has two row header then it can be combine into one row header, seperated by comma'''
        if comma_separator == True:
            connector=','
        else:
            connector=''
        string = connector.join(map(str, tuple))
        return string

    def clc_time(self,df,data_rate):
        '''Based on dataframe length, and data rate calculate time of measurement'''
        time = len(df) / data_rate
        return time

    def clc_total_weight(self,df,idx_df_weight):
        '''Calculate total weight during single measurement
        Input:
        idx_df_weight - index of weight column in dataframe
        '''
        total_weight = df.iloc[-1,idx_df_weight] - df.iloc[0,idx_df_weight]
        return total_weight

    def clc_total_mfr(self,df,idx_df_weight,data_rate):
        '''Calculate total mass flow rate during single measurement'''
        mass_flow_rate = self.clc_total_weight(df,idx_df_weight) / self.clc_time(df,data_rate)
        return mass_flow_rate

    def excel_cols(self):
        ''''Generate Excel-like column list, eg. A,B,C..AA,AB..ect.
        In order to generate the list use following code:
        list(itertools.islice(mass_flow_rate_analysis.excel_cols(), number_of_elements))
        number_of_elements - length of list
        '''
        n = 1
        while True:
            yield from (''.join(group) for group in itertools.product(string.ascii_uppercase, repeat=n))
            n += 1

    def read_data(self,path,header_index=[0,1],initial_drop_rows=10):
        '''Read data from file and convert it into DataFrame'''
        df=pd.read_csv(path,sep='\t',skiprows=1,header=header_index)
        if initial_drop_rows > 0:
            df.drop(df.index[0:initial_drop_rows],inplace=True) # first few rows are incorrect, in most cases
        df.reset_index(inplace=True,drop=True)
        orignal_header = df.columns # save orignal header
        # New header
        if len(header_index) > 1:
            new_header=list()
            for line in orignal_header:
                new_header.append(self.join_tuple(line))
            df.columns=new_header
        return df,orignal_header

    def clc_avr_window(self, df, idx_df_weight, idx_df_pressure, window_size,data_rate):
        '''Periodic mean of the data based on given period (window size). 
        Calculate a mean from window and move to following. Windows don't overlapped each other.
        Input:
        df-Dataframe
        idx_df_weight-index of weight data column in dataframe
        idx_df_pressure-index of pressure data column in dataframe
        window_size-window size to average
        data_rate-data rate of gathered data
        '''
        window_start,window_end=0,0
        window_time=window_size/data_rate
        nb_full_periods=len(df)//window_size # Only full window can be averaged.

        data_temp=list()
        weight_avr_arr=list()
        mass_flow_rate_arr=list()
        pressure_avr_arr=list()
        nb=list()

        for x in range(nb_full_periods):
            window_start=window_end
            window_end+=window_size
            weight_avr = df.iloc[window_end-1,idx_df_weight] - df.iloc[window_start,idx_df_weight]
            weight_avr_arr.append(weight_avr)
            mass_flow_rate_arr.append(weight_avr/window_time)
            pressure_avr_arr.append(df.iloc[window_start:window_end,idx_df_pressure].mean())
            nb.append(x+1)
        
        data_arr=np.concatenate((np.array(nb)[np.newaxis, :],np.array(weight_avr_arr)[np.newaxis, :],np.array(mass_flow_rate_arr)[np.newaxis, :],
                             np.array(pressure_avr_arr)[np.newaxis, :],np.array(pressure_avr_arr)[np.newaxis, :]/100000))
        # Dataframe with averaged data
        data_df=pd.DataFrame(data_arr).transpose()
        data_df.columns=['Nb.','Weight-avr [g]','Mass flow rate - avr [g/s]','Pressure - avr [Pa]','Pressure - avr [bar]']
        # Summary of whole Dataframe
        data_stats=[data_rate,window_time,window_size,np.mean(mass_flow_rate_arr),np.ptp(mass_flow_rate_arr),np.mean(pressure_avr_arr),np.ptp(pressure_avr_arr)]
        data_stats_df=pd.DataFrame(data_stats).transpose()
        data_stats_df.columns=['Data rate [Hz]','Window time [s]', 'Samples per window [-]', 'Mass flow rate - avr [g/s]', 'Range of MFR [g/s]',
        'Pressure - avr [Pa]', 'Range of Pressure [Pa]']
        return data_stats_df,data_df
    
    def data_arr(self, dir,idx_df_weight,idx_df_pressure,window_size):
        '''Using the functions defined above, read measurement case directory (one or more) and load measurement files.
        Next, by means of window function calculate periodic mean for each file and add data into aggregation array (data_case)
        then into data_all-array which is intended for all cases. 
        The function returns folowing data:
        data_case - Contains averaged data from files within one measurement case
        data_stat_case - Contains basic statistics from single measurement
        data_all - Aggregation of all measurement cases
        data_stat_all - Aggregation of statistics for all measurement cases
        mass_flow_avr - Mean mass flow rate calculated for each single measurement wihin measurement case, then aggregated into one array for all cases
        pressure_avr - Mean pressure calculated for each single measurement wihin measurement case, then aggregated into one array for all cases
        cases_list - List of all measurement case names
        data_df_header - Header which fits to data_case and data_all, intended to build pandas Dataframe
        data_stats_header  - Header which fits to data_stat_case and data_stat_all, intended to build pandas Dataframe
        '''
        data_all=list()
        mass_flow_avr=list()
        pressure_avr_=list()
        cases_list = list()
        for case, direcotry in self.open_dir_list(dir).items():
            files = list()
            data_case=list()
            mass_flow_avr_case=list()
            pressure_avr_case=list()
            cases_list.append(case)
            for idx, file in enumerate(self.get_files_from_dir(direcotry)):
                files.append(file)
                full_path = os.path.join(direcotry, file)
                date,hour,data_rate = self.read_parameters_from_header(full_path)
                df,oh = self.read_data(full_path)
                mass_flow_avr_case.append(self.clc_total_mfr(df,idx_df_weight,data_rate))
                pressure_avr_case.append(df.iloc[:,idx_df_pressure].mean())
                data_stats_df,data_df = self.clc_avr_window(df,idx_df_weight,idx_df_pressure,window_size,data_rate)

                if np.any(data_case) == False: #check if array exist
                    data_case.append(data_df.values)
                    data_stat_case=data_stats_df.values
                    data_case=np.array(data_case)
                else:
                    data_case=np.concatenate((data_case,data_df.values.reshape(-1,data_df.values.shape[0],data_df.values.shape[1])))
                    data_stat_case=np.vstack((data_stat_case,data_stats_df.values))
            data_stats_header,data_df_header=data_stats_df.columns,data_df.columns
            mass_flow_avr_case = np.array(mass_flow_avr_case)
            pressure_avr_case = np.array(pressure_avr_case)
  
            if np.any(data_all) == False:
                data_all.append(data_case)
                data_all=np.array(data_all)
                data_stat_all = data_stat_case.reshape(-1,data_stat_case.shape[0],data_stat_case.shape[1])
                mass_flow_avr = mass_flow_avr_case.reshape(-1,mass_flow_avr_case.shape[0])
                pressure_avr = pressure_avr_case.reshape(-1,pressure_avr_case.shape[0])
            else:
                data_all=np.concatenate((data_all,data_case.reshape(-1,data_case.shape[0],data_case.shape[1],data_case.shape[2])))
                data_stat_all=np.concatenate((data_stat_all,data_stat_case.reshape(-1,data_stat_case.shape[0],data_stat_case.shape[1])))
                mass_flow_avr=np.vstack((mass_flow_avr,mass_flow_avr_case.reshape(-1,mass_flow_avr_case.shape[0])))
                pressure_avr=np.vstack((pressure_avr,pressure_avr_case.reshape(-1,pressure_avr_case.shape[0])))
        return data_all,data_stat_all,mass_flow_avr,pressure_avr,cases_list,data_case,data_stat_case,data_df_header,data_stats_header

    def draw_xl_chart(self,ws,title,series_name,
                  Xax_title_r,Xax_title_c,Yax_title_r,Yax_title_c,
                  Xvalue_c_min,Xvalue_r_min,Xvalue_r_max,
                  Yvalue_c_min,Yvalue_r_min,Yvalue_r_max,
                  marker_fill_color,marker_line_color,chart_position,
                  style=13,marker_symbol='square',marker_size=8,line_no_fill=False,
                  marker_outline_color='auto',
                  x_axis_scale_min='auto',x_axis_scale_max='auto',y_axis_scale_min='auto',y_axis_scale_max='auto'):

        '''Function intended to draw a chart in Excel base on data in worksheet
        Input:
        ws - Active Excel worksheet
        title - Title of the chart
        series_name - Name of the series in the chart
        Xax_title_r - The row number of the Excel cell from which the title of X axis can be read
        Xax_title_c - The column number of the Excel cell from which the title of X axis can be read
        Yax_title_r - The row number of the Excel cell from which the title of Y axis can be read
        Yax_title_c - The column number of the Excel cell from which the title of Y axis can be read
        Xvalue_c_min - Column number of X data
        Xvalue_r_min - Start row number of X data
        Xvalue_r_max - End row number of X data
        Yvalue_c_min - Column number of Y data
        Yvalue_r_min - Start row number of Y data
        Yvalue_r_max - End row number of Y data
        marker_fill_color - Fill color in HEX code eg. FF5733
        marker_line_color - Line color in HEX
        chart_position - Chart position in worksheet eg. B1
        '''
        horizontal_offset=10
        chart = ScatterChart()
        chart.title = title
        chart.style = style
        chart.x_axis.title = ws.cell(row=Xax_title_r,column=Xax_title_c).value
        chart.y_axis.title = ws.cell(row=Yax_title_r,column=Yax_title_c).value
        xvalues = Reference(ws, min_col=Xvalue_c_min, min_row=Xvalue_r_min, max_row=Xvalue_r_max)
        values = Reference(ws, min_col=Yvalue_c_min, min_row=Yvalue_r_min, max_row=Yvalue_r_max)
        series = Series(values, xvalues, title=series_name)
        chart.series.append(series)
        series.marker.symbol = marker_symbol
        series.marker.size = marker_size
        series.marker.graphicalProperties.solidFill = marker_fill_color # Marker filling (00AAAA)
        series.marker.graphicalProperties.line.solidFill = marker_line_color # Marker outline
        if line_no_fill == False:
            series.graphicalProperties.line.width = 20000
            series.graphicalProperties.line.solidFill = marker_fill_color
        else:
            series.graphicalProperties.line.noFill=True
        if marker_outline_color == 'auto':
            pass
        else:
            series.marker.graphicalProperties.line.solidFill=marker_outline_color
        if x_axis_scale_min == 'auto':
            pass
        else:
            chart.x_axis.scaling.min = int(x_axis_scale_min)
        if x_axis_scale_max == 'auto':
            pass
        else:
            chart.x_axis.scaling.max = int(x_axis_scale_max)
        if y_axis_scale_min == 'auto':
            pass
        else:
            chart.y_axis.scaling.min = int(y_axis_scale_min)
        if y_axis_scale_max == 'auto':
            pass
        else:
            chart.y_axis.scaling.max = int(y_axis_scale_max)
        chart.legend.position = 'b'
        chart.height = 10 
        chart.width = 15
        ws.add_chart(chart, chart_position)

    def save_excel(self,path,data_all,data_stat_all,mass_flow_avr,pressure_avr,cases_list,data_df_header,data_stats_header):
        '''
        Save array data into Excel. The workbook consist with 'main' sheet which is summary of all cases and contains mean values of pressure and mass flow rate.
        There is also separate sheet intended for each case with detailed information for all measurements within a case.
        The fuction returns        
        '''
        nb_of_charts=3 # In current version there is possibility to draw three set of detailed charts within one measurement case
        dict_chart_coord={} # Dictionary with initial position 
        arr=list()
        head_case=list()
        ####### Prepare and write data into main worksheet #########
        pressure_avr=pressure_avr/100000 # conversion of pressure from Pa to bar
        for i in range(mass_flow_avr.shape[0]):
            temp=np.column_stack((pressure_avr[i],mass_flow_avr[i]))
            head_case.append(cases_list[i])
            if np.any(arr) == False:
                arr=temp
            else:
                arr=np.column_stack((arr,temp))
        columns_header = ['Pressure, [bar]','Mass flow rate, [g/s]']
        two_row_header = pd.MultiIndex.from_product([head_case,columns_header])
        df_main=pd.DataFrame(arr,columns=two_row_header,index=[x for x in range(1,arr.shape[0]+1)]) # Dataframe to be saved in main worksheet
        
        wb = Workbook()
        ws = wb.active
        ws.title = "main"
        startcol_main=1 # index-like
        startrow_main=19 # index-like
        dict_chart_coord['main']=[startcol_main,startrow_main]
        with pd.ExcelWriter(path,
            mode='w',
            engine="openpyxl",
        ) as writer:
            df_main.to_excel(writer,index=True,index_label='File nb.', sheet_name='main', startcol=startcol_main,startrow=startrow_main)
        #################################################################

        ####### Prepare and write data into seperate case worksheet #########
        for idx,case in enumerate(cases_list):
            wb.create_sheet(case)
            ws = wb[case]          
            # temp_df_1 - summary for a case, which consists of mean pressure and mass flow rate for all single measurement within a case
            temp_df_1=pd.DataFrame(df_main.iloc[:,(0+idx*2):(2+idx*2)])
            temp_df_1.columns = temp_df_1.columns[0][1],temp_df_1.columns[1][1] # instead 2 row header only 1 line header 
            _startcol=2 
            _startrow=4
            dict_chart_coord[case]=[[_startrow,_startcol]] # Add coordinates of saved data
            with pd.ExcelWriter(path,
                mode='a',
                engine="openpyxl",
                if_sheet_exists="overlay",
            ) as writer:
                temp_df_1.to_excel(writer,index=True, sheet_name=case, startcol=_startcol,startrow=_startrow)

            # temp_df_2 - Summary for single measurement within a case
            # temp_df_3 - Averaged data from window function 
            for idy,run in enumerate(data_all[idx]):
                temp_df_2=pd.DataFrame(data_stat_all[idx][idy]).transpose()
                temp_df_2.columns=data_stats_header
                temp_df_3=pd.DataFrame(data_all[idx][idy],columns=data_df_header)
                if idy == 0:
                    temp_data_4 = run
                else:
                    temp_data_4 = np.vstack((temp_data_4,run))
                _startcol=0+(len(temp_df_2.columns)+3)*idy
                _startrow=len(temp_df_1)+13+nb_of_charts*19+3
                with pd.ExcelWriter(path,
                    mode='a',
                    engine="openpyxl",
                    if_sheet_exists="overlay",
                ) as writer:
                    temp_df_2.to_excel(writer,index=False, sheet_name=case, startcol=_startcol,startrow=_startrow)
                    temp_df_3.to_excel(writer,index=False, sheet_name=case, startcol=_startcol,startrow=_startrow+2)
                if idy == 0: # Only initial coordinates are needed
                    dict_chart_coord[case].append([_startrow,_startcol])

            # temp_df_4 - Aggregation of all temp_df_3 data in one dataframe
            temp_df_4=pd.DataFrame(temp_data_4,columns=data_df_header)
            _startcol=0
            _startrow=len(temp_df_1)+13+nb_of_charts*19+3+len(temp_df_2)+len(temp_df_3)+5
            dict_chart_coord[case].append([_startrow,_startcol])
            with pd.ExcelWriter(path,
                mode='a',
                engine="openpyxl",
                if_sheet_exists="overlay",
            ) as writer:
                temp_df_4.to_excel(writer,index=True, sheet_name=case, startcol=_startcol,startrow=_startrow)
        return df_main,temp_df_1,temp_df_3,temp_df_4,dict_chart_coord
    
    def add_chart(self,path,temp_df_1,temp_df_3,temp_df_4,dict_chart_coord):
        '''
        Based on provided Excel workbook and data draw charts.
        The function draw summary chart in main worksheet and set of charts in every sheet intended for separated case.
        '''
        colors=['e6194B', '3cb44b', 'ffe119', '4363d8', 'f58231', '911eb4', '42d4f4', 'f032e6', 'bfef45', 'fabed4', '469990',
                'dcbeff', '9A6324', 'fffac8', '800000', 'aaffc3', '808000', 'ffd8b1', '000075', 'a9a9a9', 'ffffff'] # In current version number of cases is limited to 21
        excel_columns=list(itertools.islice(self.excel_cols(), 80))
        horizontal_offset=10
        wb=load_workbook(path)
        sheetnames=wb.sheetnames
        for idx,sheet in enumerate(sheetnames):
            wb.active=wb[sheet]
            ws = wb.active
            if idx == 0:
                startcol_main=1 # index-like
                startrow_main=19
                wb.active = wb[sheet]
                ws = wb.active
                chart = ScatterChart()
                chart.title = 'Mass flow rate'
                chart.style = 13
                _row=startrow_main+2
                _x_column=startcol_main+2 #pressure
                _y_column=startcol_main+3 #mass flow
                chart.x_axis.title = ws.cell(row=_row,column=_x_column).value #pressure
                chart.y_axis.title = ws.cell(row=_row,column=_y_column).value #mass flow
                chart.x_axis.scaling.min = 1
                _row=startrow_main+4
                for idw in range(len(sheetnames)-1):
                    xvalues = Reference(ws, min_col=_x_column+idw*2, min_row=_row, max_row=_row+len(temp_df_1)-1) #pressure
                    values = Reference(ws, min_col=_y_column+idw*2, min_row=_row, max_row=_row+len(temp_df_1)-1) #mass flow
                    series = Series(values, xvalues, title_from_data=False,title=sheetnames[idw+1])
                    chart.series.append(series)

                    series.marker.symbol = "square"
                    series.marker.size = 8
                    series.marker.graphicalProperties.solidFill = colors[idw] # Marker filling
                    series.marker.graphicalProperties.line.solidFill = colors[idw] # Marker outline
                    series.graphicalProperties.line.width = 20000
                    series.graphicalProperties.line.solidFill = colors[idw]

                chart.legend.position = 'b'
                chart.height = 10 
                chart.width = 15 
                ws.add_chart(chart, "B1")

                #### make_beautiful ####
                for idxx in range(len(sheetnames)*2+10): #aprox. column range
                    ws.column_dimensions[excel_columns[idxx]].width = 12
                    for ad in range(startrow_main+1,startrow_main+3):
                        _adress=excel_columns[idxx]+str(ad)
                        _cell=ws[_adress]
                        _cell.alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)

            else:
                _mass_flow_column,_pressure_column=dict_chart_coord[sheet][0][1]+2,dict_chart_coord[sheet][0][0]+1
                _title_row,_data_row=dict_chart_coord[sheet][0][0]+1,dict_chart_coord[sheet][0][0]+2
                _Xax_title_r,_Xax_title_c,_Yax_title_r,_Yax_title_c=_title_row,_mass_flow_column,_title_row,_pressure_column
                _Xvalue_c_min,_Xvalue_r_min,_Xvalue_r_max=_mass_flow_column,_data_row,_data_row+6-1
                _Yvalue_c_min,_Yvalue_r_min,_Yvalue_r_max=_pressure_column,_Xvalue_r_min,_Xvalue_r_max
                self.draw_xl_chart(ws,'MFR','def',
                                _Xax_title_r,_Xax_title_c,_Yax_title_r,_Yax_title_c,
                                _Xvalue_c_min,_Xvalue_r_min,_Xvalue_r_max,
                                _Yvalue_c_min,_Yvalue_r_min,_Yvalue_r_max,
                                colors[idx-1],colors[idx-1],'G1',x_axis_scale_min=1)
                #### make_beautiful ####
                for idyy in range(3,5):
                    _adress=excel_columns[idyy]+str(5)
                    _cell=ws[_adress]
                    _cell.alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)
                
                _pressure_column,_mass_flow_column=dict_chart_coord[sheet][2][1]+6,dict_chart_coord[sheet][2][1]+4
                _title_row,_data_row=dict_chart_coord[sheet][2][0]+1,dict_chart_coord[sheet][2][0]+2
                _Xax_title_r,_Xax_title_c=_title_row,_pressure_column
                _Yax_title_r,_Yax_title_c=_title_row,_mass_flow_column
                _Xvalue_c_min,_Xvalue_r_min,_Xvalue_r_max=_Xax_title_c,_data_row,_data_row + len(temp_df_4) -1
                _Yvalue_c_min,_Yvalue_r_min,_Yvalue_r_max=_Yax_title_c,_data_row,_data_row + len(temp_df_4) -1
                _series_name='points'
                self.draw_xl_chart(ws,'Mass flow rate - points',_series_name,
                                   _Xax_title_r,_Xax_title_c,_Yax_title_r,_Yax_title_c,
                                   _Xvalue_c_min,_Xvalue_r_min,_Xvalue_r_max,
                                   _Yvalue_c_min,_Yvalue_r_min,_Yvalue_r_max,
                                   colors[idx-1],colors[idx-1],'P1',line_no_fill=True,
                                   style=5,marker_symbol='circle',marker_size=5,
                                   marker_outline_color='000000')
                initial_vertical_step=19
                _index_row=dict_chart_coord[sheet][1][0] 
                for idz in range(len(temp_df_1)):
                    adress1=excel_columns[idz*10]+str(initial_vertical_step+1)
                    adress2=excel_columns[idz*10]+str(2*initial_vertical_step+1)
                    adress3=excel_columns[idz*10]+str(3*initial_vertical_step+1)

                    ######## Pressure - Mass flow rate ###############
                    _mass_flow_column,_pressure_column=3,5
                    _Xax_title_r,_Xax_title_c=_index_row+3,_pressure_column+idz*horizontal_offset
                    _Yax_title_r,_Yax_title_c=_Xax_title_r,_mass_flow_column+idz*horizontal_offset
                    _Xvalue_c_min,_Xvalue_r_min,_Xvalue_r_max=_Xax_title_c,_index_row+4,_index_row+4+len(temp_df_3)-1
                    _Yvalue_c_min,_Yvalue_r_min,_Yvalue_r_max=_Yax_title_c,_Xvalue_r_min,_Xvalue_r_max
                    _series_name='P='+str(int(round(ws.cell(row=_Yax_title_r+1,column=_Xax_title_c).value,1)))+'bar'
                    self.draw_xl_chart(ws,'Pressure - Mass flow rate',_series_name,
                                       _Xax_title_r,_Xax_title_c,_Yax_title_r,_Yax_title_c,
                                       _Xvalue_c_min,_Xvalue_r_min,_Xvalue_r_max,
                                       _Yvalue_c_min,_Yvalue_r_min,_Yvalue_r_max,
                                       colors[idx-1],colors[idx-1],adress1,line_no_fill=True)
                    ######################################

                    ######## Mass flow rate per case ###############
                    _Nb_column=1
                    _Xax_title_r,_Xax_title_c=_index_row+3,_Nb_column+idz*horizontal_offset
                    _Yax_title_r,_Yax_title_c=_Xax_title_r,_mass_flow_column+idz*horizontal_offset

                    _Xvalue_c_min=_Xax_title_c
                    _Xvalue_r_min=_index_row+4
                    _Xvalue_r_max=_index_row+4+len(temp_df_3)-1

                    _Yvalue_c_min=_Yax_title_c
                    _Yvalue_r_min=_Xvalue_r_min
                    _Yvalue_r_max=_Xvalue_r_max

                    _series_name='P='+str(int(round(ws.cell(row=_Yax_title_r+1,column=_pressure_column+idz*horizontal_offset).value,1)))+'bar'
                    
                    self.draw_xl_chart(ws,'Mass flow rate per case',_series_name,
                            _Xax_title_r,_Xax_title_c,_Yax_title_r,_Yax_title_c,
                            _Xvalue_c_min,_Xvalue_r_min,_Xvalue_r_max,
                            _Yvalue_c_min,_Yvalue_r_min,_Yvalue_r_max,
                            colors[idx-1],colors[idx-1],adress2,line_no_fill=True)
                    ######################################

                    ######## Pressure - Mass flow rate ###############
                    _Xax_title_r=_index_row+3
                    _Xax_title_c=_Nb_column+idz*horizontal_offset
                    _Yax_title_r=_Xax_title_r
                    _Yax_title_c=_pressure_column+idz*horizontal_offset

                    _Xvalue_c_min=_Xax_title_c
                    _Xvalue_r_min=_index_row+4
                    _Xvalue_r_max=_index_row+4+len(temp_df_3)-1

                    _Yvalue_c_min=_Yax_title_c
                    _Yvalue_r_min=_Xvalue_r_min
                    _Yvalue_r_max=_Xvalue_r_max

                    _series_name='P='+str(int(round(ws.cell(row=_Yax_title_r+1,column=_pressure_column+idz*horizontal_offset).value,1)))+'bar'

                    self.draw_xl_chart(ws,'Absolute pressure',_series_name,
                            _Xax_title_r,_Xax_title_c,_Yax_title_r,_Yax_title_c,
                            _Xvalue_c_min,_Xvalue_r_min,_Xvalue_r_max,
                            _Yvalue_c_min,_Yvalue_r_min,_Yvalue_r_max,
                            colors[idx-1],colors[idx-1],adress3,line_no_fill=True)                 
                    ######################################

                ### make_beautiful ####
        for idzz in range(len(temp_df_1)*10):
            ws.column_dimensions[excel_columns[idzz]].width = 10
            _adress=excel_columns[idzz]+str(dict_chart_coord[sheet][1][0]+1)
            _cell=ws[_adress]
            _cell.alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)
            _adress=excel_columns[idzz]+str(dict_chart_coord[sheet][1][0]+3)
            _cell=ws[_adress]
            _cell.alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)
            _adress=excel_columns[idzz]+str(dict_chart_coord[sheet][2][0]+1)
            _cell=ws[_adress]
            _cell.alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)
        wb.save(path)